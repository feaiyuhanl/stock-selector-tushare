"""
缓存管理模块：管理本地Excel缓存
"""
import pandas as pd
import os
import glob
import threading
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import config
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle


class CacheManager:
    """缓存管理器"""
    
    def __init__(self, cache_dir: str = "cache"):
        """
        初始化缓存管理器
        Args:
            cache_dir: 缓存目录
        """
        self.cache_dir = cache_dir
        self._ensure_cache_dir()
        
        # 统一的缓存目录结构（扁平化，只保留1级目录）
        # K线数据目录（所有K线数据统一放在这里，包括股票、板块、概念的K线）
        self.kline_cache_dir = os.path.join(cache_dir, "kline")
        
        # 元数据目录：存储汇总信息和集中存储文件（用于快速查询和兼容性）
        self.meta_dir = os.path.join(cache_dir, "meta")
        
        # 集中存储文件路径（放在meta目录下，统一管理）
        # 这些文件用于快速批量查询，但主要使用独立文件存储
        self.fundamental_cache = os.path.join(self.meta_dir, "fundamental_data.xlsx")
        self.financial_cache = os.path.join(self.meta_dir, "financial_data.xlsx")
        self.sectors_cache = os.path.join(self.meta_dir, "stock_sectors.xlsx")
        self.concepts_cache = os.path.join(self.meta_dir, "stock_concepts.xlsx")
        self.stock_list_cache = os.path.join(self.meta_dir, "stock_list.xlsx")
        
        # 确保所有缓存目录存在
        required_dirs = [
            self.kline_cache_dir,
            self.meta_dir
        ]
        for cache_dir_path in required_dirs:
            if not os.path.exists(cache_dir_path):
                os.makedirs(cache_dir_path)
        
        # 清理可能存在的临时文件
        self._cleanup_temp_files()
        
        # 缓存有效期（天数）
        # 说明：超过有效期后，缓存自动失效，会重新从API获取数据
        self.cache_valid_days = {
            'fundamental': 7,      # 缓存有效期7天（超过7天自动失效）
            'financial': 7,        # 缓存有效期7天（超过7天自动失效）
            'sectors': 7,          # 缓存有效期7天（超过7天自动失效）
            'concepts': 7,         # 缓存有效期7天（超过7天自动失效）
            'stock_list': 1,       # 缓存有效期1天（超过1天自动失效）
            'kline': 1,            # 缓存有效期1天（必须是今天，否则失效）
        }
        
        # 定义哪些数据类型是低频的（变化频率低，但缓存失效后仍会重新获取）
        self.low_frequency_types = ['financial', 'sectors', 'concepts']
        
        # 文件级别的锁，防止并发写入导致文件句柄泄露（Windows上特别重要）
        self._file_locks = {
            'fundamental': threading.Lock(),
            'financial': threading.Lock(),
            'sectors': threading.Lock(),
            'concepts': threading.Lock(),
            'stock_list': threading.Lock(),
        }
    
    def _ensure_cache_dir(self):
        """确保缓存目录存在"""
        if not os.path.exists(self.cache_dir):
            os.makedirs(self.cache_dir)
    
    def _cleanup_temp_files(self):
        """清理可能存在的临时文件（程序异常退出时可能留下）"""
        try:
            # 清理meta目录下的临时文件
            if os.path.exists(self.meta_dir):
                for filename in os.listdir(self.meta_dir):
                    # 清理旧的.tmp文件和新的_temp.xlsx临时文件
                    if filename.endswith('.tmp') or filename.endswith('_temp.xlsx'):
                        temp_file = os.path.join(self.meta_dir, filename)
                        try:
                            os.remove(temp_file)
                        except:
                            pass
        except:
            pass
    
    def _is_cache_valid(self, cache_file: str, cache_type: str) -> bool:
        """
        检查缓存是否有效
        Args:
            cache_file: 缓存文件路径
            cache_type: 缓存类型
        Returns:
            是否有效
        """
        if not os.path.exists(cache_file):
            return False
        
        # 检查文件修改时间
        mtime = datetime.fromtimestamp(os.path.getmtime(cache_file))
        valid_days = self.cache_valid_days.get(cache_type, 7)
        
        # 对于当日缓存（K线数据），检查是否是同一天
        if cache_type == 'kline':
            return mtime.date() == datetime.now().date()
        
        return (datetime.now() - mtime).days < valid_days
    
    def _is_excel_file_valid(self, file_path: str) -> bool:
        """
        检查Excel文件是否有效（未损坏）
        Args:
            file_path: Excel文件路径
        Returns:
            文件是否有效
        """
        if not os.path.exists(file_path):
            return False
        
        try:
            # 尝试读取文件，检查是否损坏
            df = pd.read_excel(file_path, engine='openpyxl', nrows=1)
            return True
        except Exception as e:
            # 文件损坏，返回False
            return False
    
    def _repair_corrupted_cache(self, cache_file: str, cache_type: str = 'fundamental'):
        """
        修复损坏的缓存文件
        Args:
            cache_file: 缓存文件路径
            cache_type: 缓存类型
        """
        try:
            # 对于sectors和concepts类型，如果文件损坏，直接删除，让程序重新创建
            if cache_type in ['sectors', 'concepts']:
                try:
                    # 备份损坏的文件
                    backup_file = cache_file.replace('.xlsx', '_backup.xlsx')
                    if os.path.exists(cache_file):
                        try:
                            os.rename(cache_file, backup_file)
                        except:
                            # 如果重命名失败，直接删除
                            try:
                                os.remove(cache_file)
                            except:
                                pass
                except:
                    pass
                return
            
            # 已迁移到meta目录，不再使用独立文件恢复
            # 如果文件损坏，直接删除，让程序重新创建
            try:
                if os.path.exists(cache_file):
                    os.remove(cache_file)
                    print(f"已删除损坏的{cache_type}缓存文件（将重新创建）")
            except:
                pass
        except Exception as e:
            # 修复失败，删除损坏文件
            try:
                if os.path.exists(cache_file):
                    os.remove(cache_file)
                    print(f"修复失败，已删除损坏的{cache_type}缓存文件: {e}")
            except:
                pass
    
    def _is_data_valid(self, data: Dict, data_type: str) -> bool:
        """
        检查数据是否有效（关键指标不为空）
        优化：区分0值（可能是正常值）和None值（数据缺失）
        Args:
            data: 数据字典
            data_type: 数据类型 ('fundamental', 'financial')
        Returns:
            是否有效
        """
        if not data or not isinstance(data, dict):
            return False
        
        if data_type == 'fundamental':
            # 基本面数据：至少需要pe_ratio或pb_ratio之一有效（不为None）
            # 改进：区分0值（可能是正常值，如亏损股的PE=0）和None值（数据缺失）
            pe = data.get('pe_ratio')
            pb = data.get('pb_ratio')
            
            # 如果两个都是None，视为无效（数据缺失）
            # 如果至少有一个不是None（即使是0），视为有效
            if pe is None and pb is None:
                return False
            return True
        elif data_type == 'financial':
            # 财务数据：至少需要roe有效（不为None）
            roe = data.get('roe')
            if roe is None:
                return False
            return True
        
        return True
    
    def get_fundamental(self, stock_code: str, force_refresh: bool = False) -> Optional[Dict]:
        """
        从缓存获取基本面数据（优先从集中文件读取，兼容独立文件）
        Args:
            stock_code: 股票代码
            force_refresh: 是否强制刷新
        Returns:
            基本面数据字典
        """
        if force_refresh:
            return None
        
        # 确保代码格式化为6位字符串
        stock_code = str(stock_code).zfill(6)
        
        # 方式1：优先从集中文件读取（主要方式）
        if self._is_cache_valid(self.fundamental_cache, 'fundamental'):
            # 先检查文件是否损坏
            if not self._is_excel_file_valid(self.fundamental_cache):
                # 文件损坏，尝试修复
                self._repair_corrupted_cache(self.fundamental_cache, 'fundamental')
                # 修复后如果文件仍不存在，跳过
                if not os.path.exists(self.fundamental_cache):
                    pass
                else:
                    # 修复后再次尝试读取
                    try:
                        df = pd.read_excel(self.fundamental_cache, engine='openpyxl')
                        # 确保code列是字符串格式
                        if 'code' in df.columns:
                            df['code'] = df['code'].astype(str).str.zfill(6)
                        stock_data = df[df['code'] == stock_code]
                        if not stock_data.empty:
                            data = stock_data.iloc[0].to_dict()
                            # 验证数据有效性
                            if self._is_data_valid(data, 'fundamental'):
                                return data
                    except Exception as e:
                        # 修复后仍然失败，删除文件
                        try:
                            os.remove(self.fundamental_cache)
                        except:
                            pass
            else:
                # 文件有效，正常读取
                try:
                    df = pd.read_excel(self.fundamental_cache, engine='openpyxl')
                    # 确保code列是字符串格式
                    if 'code' in df.columns:
                        df['code'] = df['code'].astype(str).str.zfill(6)
                    stock_data = df[df['code'] == stock_code]
                    if not stock_data.empty:
                        data = stock_data.iloc[0].to_dict()
                        # 验证数据有效性
                        if self._is_data_valid(data, 'fundamental'):
                            return data
                except Exception as e:
                    # 读取失败，可能是文件损坏，尝试修复
                    print(f"读取基本面集中缓存失败: {e}")
                    self._repair_corrupted_cache(self.fundamental_cache, 'fundamental')
        
        # 方式2：从独立文件读取（兼容旧数据，已迁移到meta目录，不再使用）
        # 旧数据已迁移到meta目录，不再读取独立文件
        
        return None
    
    def save_fundamental(self, stock_code: str, data: Dict):
        """
        保存基本面数据到缓存（优化策略：只有一行数据的合并到集中文件，多行数据才用独立文件）
        Args:
            stock_code: 股票代码
            data: 基本面数据
        """
        # 使用文件锁防止并发写入
        with self._file_locks['fundamental']:
            try:
                # 验证数据有效性
                if not self._is_data_valid(data, 'fundamental'):
                    return  # 无效数据不保存
                
                data_to_save = data.copy()
                data_to_save['code'] = stock_code
                data_to_save['update_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # 基本面数据只有一行，直接保存到集中文件（不创建独立文件）
                # 先检查文件是否损坏
                if os.path.exists(self.fundamental_cache):
                    if not self._is_excel_file_valid(self.fundamental_cache):
                        # 文件损坏，尝试修复
                        self._repair_corrupted_cache(self.fundamental_cache, 'fundamental')
                        # 如果修复失败，创建新文件
                        if not os.path.exists(self.fundamental_cache) or not self._is_excel_file_valid(self.fundamental_cache):
                            df = pd.DataFrame()
                        else:
                            df = pd.read_excel(self.fundamental_cache, engine='openpyxl')
                    else:
                        df = pd.read_excel(self.fundamental_cache, engine='openpyxl')
                    # 确保code列是字符串格式
                    if 'code' in df.columns:
                        df['code'] = df['code'].astype(str).str.zfill(6)
                    # 移除旧数据
                    df = df[df['code'] != stock_code]
                else:
                    df = pd.DataFrame()
                
                new_row = pd.DataFrame([data_to_save])
                df = pd.concat([df, new_row], ignore_index=True)
                
                # 使用临时文件写入，避免写入过程中文件损坏
                # 使用.xlsx扩展名，pandas才能正确识别文件类型
                temp_file = self.fundamental_cache.replace('.xlsx', '_temp.xlsx')
                try:
                    # 写入临时文件
                    df.to_excel(temp_file, index=False, engine='openpyxl')
                    # 确保写入完成，文件句柄关闭
                    del df
                    time.sleep(0.1)  # 短暂延迟，确保文件句柄释放
                    
                    # 删除原文件（带重试机制）
                    max_retries = 5
                    for attempt in range(max_retries):
                        try:
                            if os.path.exists(self.fundamental_cache):
                                os.remove(self.fundamental_cache)
                            break
                        except PermissionError:
                            if attempt < max_retries - 1:
                                time.sleep(0.2)  # 等待文件句柄释放
                            else:
                                raise
                    
                    # 重命名临时文件
                    os.rename(temp_file, self.fundamental_cache)
                except Exception as e:
                    # 如果写入失败，删除临时文件
                    if os.path.exists(temp_file):
                        try:
                            os.remove(temp_file)
                        except:
                            pass
                    raise e
                
                # 旧数据已迁移到meta目录，不再使用独立文件
            except Exception as e:
                print(f"保存基本面缓存失败: {e}")
    
    def batch_save_fundamental(self, data_dict: Dict[str, Dict]):
        """
        批量保存基本面数据（提高效率，减少IO操作，只保存到集中文件）
        Args:
            data_dict: {stock_code: data_dict} 字典
        """
        if not data_dict:
            return
        
        # 使用文件锁防止并发写入
        with self._file_locks['fundamental']:
            try:
                # 读取现有集中文件
                df = None
                if os.path.exists(self.fundamental_cache):
                    # 检查文件是否损坏
                    if not self._is_excel_file_valid(self.fundamental_cache):
                        # 文件损坏，尝试修复
                        self._repair_corrupted_cache(self.fundamental_cache, 'fundamental')
                        # 如果修复失败，创建新文件
                        if not os.path.exists(self.fundamental_cache) or not self._is_excel_file_valid(self.fundamental_cache):
                            df = pd.DataFrame()
                        else:
                            df = pd.read_excel(self.fundamental_cache, engine='openpyxl')
                    else:
                        df = pd.read_excel(self.fundamental_cache, engine='openpyxl')
                    # 确保code列是字符串格式
                    if 'code' in df.columns:
                        df['code'] = df['code'].astype(str).str.zfill(6)
                    # 移除要更新的股票旧数据
                    existing_codes = set(str(code).zfill(6) for code in data_dict.keys())
                    df = df[~df['code'].isin(existing_codes)]
                else:
                    df = pd.DataFrame()
                
                # 准备新数据（只保存到集中文件，不创建独立文件）
                new_rows = []
                for stock_code, data in data_dict.items():
                    # 验证数据有效性
                    if not self._is_data_valid(data, 'fundamental'):
                        continue
                    
                    data_to_save = data.copy()
                    data_to_save['code'] = str(stock_code).zfill(6)
                    data_to_save['update_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    new_rows.append(data_to_save)
                
                # 批量更新集中文件（使用临时文件，避免写入过程中文件损坏）
                if new_rows:
                    new_df = pd.DataFrame(new_rows)
                    df = pd.concat([df, new_df], ignore_index=True)
                    # 使用.xlsx扩展名，pandas才能正确识别文件类型
                    temp_file = self.fundamental_cache.replace('.xlsx', '_temp.xlsx')
                    try:
                        # 写入临时文件
                        df.to_excel(temp_file, index=False, engine='openpyxl')
                        # 确保写入完成，文件句柄关闭
                        del df
                        time.sleep(0.1)  # 短暂延迟，确保文件句柄释放
                        
                        # 删除原文件（带重试机制）
                        max_retries = 5
                        for attempt in range(max_retries):
                            try:
                                if os.path.exists(self.fundamental_cache):
                                    os.remove(self.fundamental_cache)
                                break
                            except PermissionError:
                                if attempt < max_retries - 1:
                                    time.sleep(0.2)  # 等待文件句柄释放
                                else:
                                    raise
                        
                        # 重命名临时文件
                        os.rename(temp_file, self.fundamental_cache)
                    except Exception as e:
                        # 如果写入失败，删除临时文件
                        if os.path.exists(temp_file):
                            try:
                                os.remove(temp_file)
                            except:
                                pass
                        raise e
            except Exception as e:
                print(f"批量保存基本面缓存失败: {e}")
    
    def get_financial(self, stock_code: str, force_refresh: bool = False) -> Optional[Dict]:
        """
        从缓存获取财务数据（优先从集中文件读取，兼容独立文件）
        Args:
            stock_code: 股票代码
            force_refresh: 是否强制刷新
        Returns:
            财务数据字典
        """
        if force_refresh:
            return None
        
        # 确保代码格式化为6位字符串
        stock_code = str(stock_code).zfill(6)
        
        # 方式1：优先从集中文件读取（主要方式）
        if self._is_cache_valid(self.financial_cache, 'financial'):
            # 先检查文件是否损坏
            if not self._is_excel_file_valid(self.financial_cache):
                # 文件损坏，尝试修复
                self._repair_corrupted_cache(self.financial_cache, 'financial')
                # 修复后如果文件仍不存在，跳过
                if not os.path.exists(self.financial_cache):
                    pass
                else:
                    # 修复后再次尝试读取
                    try:
                        df = pd.read_excel(self.financial_cache, engine='openpyxl')
                        # 确保code列是字符串格式
                        if 'code' in df.columns:
                            df['code'] = df['code'].astype(str).str.zfill(6)
                        stock_data = df[df['code'] == stock_code]
                        if not stock_data.empty:
                            data = stock_data.iloc[0].to_dict()
                            # 验证数据有效性
                            if self._is_data_valid(data, 'financial'):
                                return data
                    except Exception as e:
                        # 修复后仍然失败，删除文件
                        try:
                            os.remove(self.financial_cache)
                        except:
                            pass
            else:
                # 文件有效，正常读取
                try:
                    df = pd.read_excel(self.financial_cache, engine='openpyxl')
                    # 确保code列是字符串格式
                    if 'code' in df.columns:
                        df['code'] = df['code'].astype(str).str.zfill(6)
                    stock_data = df[df['code'] == stock_code]
                    if not stock_data.empty:
                        data = stock_data.iloc[0].to_dict()
                        # 验证数据有效性
                        if self._is_data_valid(data, 'financial'):
                            return data
                except Exception as e:
                    # 读取失败，可能是文件损坏，尝试修复
                    print(f"读取财务集中缓存失败: {e}")
                    self._repair_corrupted_cache(self.financial_cache, 'financial')
        
        # 方式2：从独立文件读取（兼容旧数据，已迁移到meta目录，不再使用）
        # 旧数据已迁移到meta目录，不再读取独立文件
        
        return None
    
    def save_financial(self, stock_code: str, data: Dict):
        """
        保存财务数据到缓存（优化策略：只有一行数据的合并到集中文件，多行数据才用独立文件）
        Args:
            stock_code: 股票代码
            data: 财务数据
        """
        # 使用文件锁防止并发写入
        with self._file_locks['financial']:
            try:
                # 验证数据有效性
                if not self._is_data_valid(data, 'financial'):
                    return  # 无效数据不保存
                
                data_to_save = data.copy()
                data_to_save['code'] = stock_code
                data_to_save['update_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # 财务数据只有一行，直接保存到集中文件（不创建独立文件）
                # 先检查文件是否损坏
                if os.path.exists(self.financial_cache):
                    if not self._is_excel_file_valid(self.financial_cache):
                        # 文件损坏，尝试修复
                        self._repair_corrupted_cache(self.financial_cache, 'financial')
                        # 如果修复失败，创建新文件
                        if not os.path.exists(self.financial_cache) or not self._is_excel_file_valid(self.financial_cache):
                            df = pd.DataFrame()
                        else:
                            df = pd.read_excel(self.financial_cache, engine='openpyxl')
                    else:
                        df = pd.read_excel(self.financial_cache, engine='openpyxl')
                    # 确保code列是字符串格式
                    if 'code' in df.columns:
                        df['code'] = df['code'].astype(str).str.zfill(6)
                    # 移除旧数据
                    df = df[df['code'] != stock_code]
                else:
                    df = pd.DataFrame()
                
                new_row = pd.DataFrame([data_to_save])
                df = pd.concat([df, new_row], ignore_index=True)
                
                # 使用临时文件写入，避免写入过程中文件损坏
                # 使用.xlsx扩展名，pandas才能正确识别文件类型
                temp_file = self.financial_cache.replace('.xlsx', '_temp.xlsx')
                try:
                    # 写入临时文件
                    df.to_excel(temp_file, index=False, engine='openpyxl')
                    # 确保写入完成，文件句柄关闭
                    del df
                    time.sleep(0.1)  # 短暂延迟，确保文件句柄释放
                    
                    # 删除原文件（带重试机制）
                    max_retries = 5
                    for attempt in range(max_retries):
                        try:
                            if os.path.exists(self.financial_cache):
                                os.remove(self.financial_cache)
                            break
                        except PermissionError:
                            if attempt < max_retries - 1:
                                time.sleep(0.2)  # 等待文件句柄释放
                            else:
                                raise
                    
                    # 重命名临时文件
                    os.rename(temp_file, self.financial_cache)
                except Exception as e:
                    # 如果写入失败，删除临时文件
                    if os.path.exists(temp_file):
                        try:
                            os.remove(temp_file)
                        except:
                            pass
                    raise e
                
                # 旧数据已迁移到meta目录，不再使用独立文件
            except Exception as e:
                print(f"保存财务缓存失败: {e}")
    
    def batch_save_financial(self, data_dict: Dict[str, Dict]):
        """
        批量保存财务数据（提高效率，减少IO操作，只保存到集中文件）
        Args:
            data_dict: {stock_code: data_dict} 字典
        """
        if not data_dict:
            return
        
        # 使用文件锁防止并发写入
        with self._file_locks['financial']:
            try:
                # 读取现有集中文件
                df = None
                if os.path.exists(self.financial_cache):
                    # 检查文件是否损坏
                    if not self._is_excel_file_valid(self.financial_cache):
                        # 文件损坏，尝试修复
                        self._repair_corrupted_cache(self.financial_cache, 'financial')
                        # 如果修复失败，创建新文件
                        if not os.path.exists(self.financial_cache) or not self._is_excel_file_valid(self.financial_cache):
                            df = pd.DataFrame()
                        else:
                            df = pd.read_excel(self.financial_cache, engine='openpyxl')
                    else:
                        df = pd.read_excel(self.financial_cache, engine='openpyxl')
                    # 确保code列是字符串格式
                    if 'code' in df.columns:
                        df['code'] = df['code'].astype(str).str.zfill(6)
                    # 移除要更新的股票旧数据
                    existing_codes = set(str(code).zfill(6) for code in data_dict.keys())
                    df = df[~df['code'].isin(existing_codes)]
                else:
                    df = pd.DataFrame()
                
                # 准备新数据（只保存到集中文件，不创建独立文件）
                new_rows = []
                for stock_code, data in data_dict.items():
                    # 验证数据有效性
                    if not self._is_data_valid(data, 'financial'):
                        continue
                    
                    data_to_save = data.copy()
                    data_to_save['code'] = str(stock_code).zfill(6)
                    data_to_save['update_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    new_rows.append(data_to_save)
                
                # 批量更新集中文件（使用临时文件，避免写入过程中文件损坏）
                if new_rows:
                    new_df = pd.DataFrame(new_rows)
                    df = pd.concat([df, new_df], ignore_index=True)
                    # 使用.xlsx扩展名，pandas才能正确识别文件类型
                    temp_file = self.financial_cache.replace('.xlsx', '_temp.xlsx')
                    try:
                        # 写入临时文件
                        df.to_excel(temp_file, index=False, engine='openpyxl')
                        # 确保写入完成，文件句柄关闭
                        del df
                        time.sleep(0.1)  # 短暂延迟，确保文件句柄释放
                        
                        # 删除原文件（带重试机制）
                        max_retries = 5
                        for attempt in range(max_retries):
                            try:
                                if os.path.exists(self.financial_cache):
                                    os.remove(self.financial_cache)
                                break
                            except PermissionError:
                                if attempt < max_retries - 1:
                                    time.sleep(0.2)  # 等待文件句柄释放
                                else:
                                    raise
                        
                        # 重命名临时文件
                        os.rename(temp_file, self.financial_cache)
                    except Exception as e:
                        # 如果写入失败，删除临时文件
                        if os.path.exists(temp_file):
                            try:
                                os.remove(temp_file)
                            except:
                                pass
                        raise e
            except Exception as e:
                print(f"批量保存财务缓存失败: {e}")
    
    def get_stock_sectors(self, stock_code: str, force_refresh: bool = False) -> Optional[List[str]]:
        """
        从缓存获取股票板块
        Args:
            stock_code: 股票代码
            force_refresh: 是否强制刷新
        Returns:
            板块列表，如果缓存中没有数据返回None，如果缓存中保存了空列表返回[]
        """
        if not force_refresh and self._is_cache_valid(self.sectors_cache, 'sectors'):
            try:
                # 检查文件是否损坏
                if not self._is_excel_file_valid(self.sectors_cache):
                    # 文件损坏，尝试修复
                    self._repair_corrupted_cache(self.sectors_cache, 'sectors')
                    # 如果修复失败，返回None
                    if not os.path.exists(self.sectors_cache) or not self._is_excel_file_valid(self.sectors_cache):
                        return None
                
                df = pd.read_excel(self.sectors_cache, engine='openpyxl')
                # 确保code列是字符串格式
                if 'code' in df.columns:
                    df['code'] = df['code'].astype(str).str.zfill(6)
                stock_data = df[df['code'] == stock_code]
                if not stock_data.empty:
                    sectors_str = stock_data.iloc[0].get('sectors', '')
                    # 如果sectors字段存在但为空字符串，说明之前保存了空列表
                    if sectors_str == '':
                        # 检查缓存时间，如果超过3天，返回None让fetcher重新获取
                        update_time_str = stock_data.iloc[0].get('update_time', '')
                        if update_time_str:
                            try:
                                update_time = datetime.strptime(update_time_str, '%Y-%m-%d %H:%M:%S')
                                days_old = (datetime.now() - update_time).days
                                if days_old > 3:  # 空列表缓存超过3天，重新获取
                                    return None
                            except:
                                pass
                        return []
                    # 如果sectors字段有值，返回解析后的列表
                    return sectors_str.split(',') if sectors_str else []
            except Exception as e:
                print(f"读取板块缓存失败: {e}")
        # 缓存中没有数据，返回None
        return None
    
    def save_stock_sectors(self, stock_code: str, sectors: List[str]):
        """
        保存股票板块到缓存
        Args:
            stock_code: 股票代码
            sectors: 板块列表（空列表也会保存，用于标记该股票确实没有板块信息）
        """
        # 使用文件锁防止并发写入
        with self._file_locks['sectors']:
            try:
                df = None
                if os.path.exists(self.sectors_cache):
                    # 检查文件是否损坏
                    if not self._is_excel_file_valid(self.sectors_cache):
                        # 文件损坏，尝试修复
                        self._repair_corrupted_cache(self.sectors_cache, 'sectors')
                        # 如果修复失败，创建新文件
                        if not os.path.exists(self.sectors_cache) or not self._is_excel_file_valid(self.sectors_cache):
                            df = pd.DataFrame()
                        else:
                            try:
                                df = pd.read_excel(self.sectors_cache, engine='openpyxl')
                            except:
                                df = pd.DataFrame()
                    else:
                        # 文件有效，正常读取
                        try:
                            df = pd.read_excel(self.sectors_cache, engine='openpyxl')
                        except Exception as e:
                            # 读取失败，可能是文件损坏，尝试修复
                            print(f"读取板块缓存失败，尝试修复: {e}")
                            self._repair_corrupted_cache(self.sectors_cache, 'sectors')
                            if os.path.exists(self.sectors_cache) and self._is_excel_file_valid(self.sectors_cache):
                                try:
                                    df = pd.read_excel(self.sectors_cache, engine='openpyxl')
                                except:
                                    df = pd.DataFrame()
                            else:
                                df = pd.DataFrame()
                    
                    if df is not None and not df.empty:
                        # 确保code列是字符串格式
                        if 'code' in df.columns:
                            df['code'] = df['code'].astype(str).str.zfill(6)
                        # 移除要更新的股票旧数据
                        df = df[df['code'] != stock_code]
                else:
                    df = pd.DataFrame()
                
                new_row = pd.DataFrame([{
                    'code': str(stock_code).zfill(6),
                    'sectors': ','.join(sectors) if sectors else '',  # 空列表保存为空字符串
                    'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }])
                df = pd.concat([df, new_row], ignore_index=True)
                
                # 确保目录存在
                os.makedirs(os.path.dirname(self.sectors_cache), exist_ok=True)
                
                # 写入临时文件，然后重命名（原子操作）
                temp_file = self.sectors_cache.replace('.xlsx', '_temp.xlsx')
                df.to_excel(temp_file, index=False, engine='openpyxl')
                
                # 删除旧文件并重命名临时文件
                if os.path.exists(self.sectors_cache):
                    try:
                        os.remove(self.sectors_cache)
                    except:
                        pass
                os.rename(temp_file, self.sectors_cache)
                
            except Exception as e:
                print(f"保存板块缓存失败: {e}")
                import traceback
                traceback.print_exc()
                # 如果保存失败，尝试删除损坏的文件
                if os.path.exists(self.sectors_cache):
                    try:
                        if not self._is_excel_file_valid(self.sectors_cache):
                            backup_file = self.sectors_cache.replace('.xlsx', '_backup.xlsx')
                            try:
                                os.rename(self.sectors_cache, backup_file)
                            except:
                                pass
                    except:
                        pass
    
    def get_stock_concepts(self, stock_code: str, force_refresh: bool = False) -> Optional[List[str]]:
        """
        从缓存获取股票概念
        Args:
            stock_code: 股票代码
            force_refresh: 是否强制刷新
        Returns:
            概念列表，如果缓存中没有数据返回None，如果缓存中保存了空列表返回[]
        """
        if not force_refresh and self._is_cache_valid(self.concepts_cache, 'concepts'):
            try:
                # 检查文件是否损坏
                if not self._is_excel_file_valid(self.concepts_cache):
                    # 文件损坏，尝试修复
                    self._repair_corrupted_cache(self.concepts_cache, 'concepts')
                    # 如果修复失败，返回None
                    if not os.path.exists(self.concepts_cache) or not self._is_excel_file_valid(self.concepts_cache):
                        return None
                
                df = pd.read_excel(self.concepts_cache, engine='openpyxl')
                # 确保code列是字符串格式
                if 'code' in df.columns:
                    df['code'] = df['code'].astype(str).str.zfill(6)
                stock_data = df[df['code'] == stock_code]
                if not stock_data.empty:
                    concepts_str = stock_data.iloc[0].get('concepts', '')
                    # 如果concepts字段存在但为空字符串，说明之前保存了空列表
                    if concepts_str == '':
                        # 检查缓存时间，如果超过3天，返回None让fetcher重新获取
                        update_time_str = stock_data.iloc[0].get('update_time', '')
                        if update_time_str:
                            try:
                                update_time = datetime.strptime(update_time_str, '%Y-%m-%d %H:%M:%S')
                                days_old = (datetime.now() - update_time).days
                                if days_old > 3:  # 空列表缓存超过3天，重新获取
                                    return None
                            except:
                                pass
                        return []
                    # 如果concepts字段有值，返回解析后的列表
                    return concepts_str.split(',') if concepts_str else []
            except Exception as e:
                print(f"读取概念缓存失败: {e}")
        # 缓存中没有数据，返回None
        return None
    
    def save_stock_concepts(self, stock_code: str, concepts: List[str]):
        """
        保存股票概念到缓存
        Args:
            stock_code: 股票代码
            concepts: 概念列表（空列表也会保存，用于标记该股票确实没有概念信息）
        """
        # 使用文件锁防止并发写入
        with self._file_locks['concepts']:
            try:
                df = None
                if os.path.exists(self.concepts_cache):
                    # 检查文件是否损坏
                    if not self._is_excel_file_valid(self.concepts_cache):
                        # 文件损坏，尝试修复
                        self._repair_corrupted_cache(self.concepts_cache, 'concepts')
                        # 如果修复失败，创建新文件
                        if not os.path.exists(self.concepts_cache) or not self._is_excel_file_valid(self.concepts_cache):
                            df = pd.DataFrame()
                        else:
                            try:
                                df = pd.read_excel(self.concepts_cache, engine='openpyxl')
                            except:
                                df = pd.DataFrame()
                    else:
                        # 文件有效，正常读取
                        try:
                            df = pd.read_excel(self.concepts_cache, engine='openpyxl')
                        except Exception as e:
                            # 读取失败，可能是文件损坏，尝试修复
                            print(f"读取概念缓存失败，尝试修复: {e}")
                            self._repair_corrupted_cache(self.concepts_cache, 'concepts')
                            if os.path.exists(self.concepts_cache) and self._is_excel_file_valid(self.concepts_cache):
                                try:
                                    df = pd.read_excel(self.concepts_cache, engine='openpyxl')
                                except:
                                    df = pd.DataFrame()
                            else:
                                df = pd.DataFrame()
                    
                    if df is not None and not df.empty:
                        # 确保code列是字符串格式
                        if 'code' in df.columns:
                            df['code'] = df['code'].astype(str).str.zfill(6)
                        # 移除要更新的股票旧数据
                        df = df[df['code'] != stock_code]
                else:
                    df = pd.DataFrame()
                
                new_row = pd.DataFrame([{
                    'code': str(stock_code).zfill(6),
                    'concepts': ','.join(concepts) if concepts else '',  # 空列表保存为空字符串
                    'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }])
                df = pd.concat([df, new_row], ignore_index=True)
                
                # 确保目录存在
                os.makedirs(os.path.dirname(self.concepts_cache), exist_ok=True)
                
                # 写入临时文件，然后重命名（原子操作）
                temp_file = self.concepts_cache.replace('.xlsx', '_temp.xlsx')
                df.to_excel(temp_file, index=False, engine='openpyxl')
                
                # 删除旧文件并重命名临时文件
                if os.path.exists(self.concepts_cache):
                    try:
                        os.remove(self.concepts_cache)
                    except:
                        pass
                os.rename(temp_file, self.concepts_cache)
                
            except Exception as e:
                print(f"保存概念缓存失败: {e}")
                import traceback
                traceback.print_exc()
                # 如果保存失败，尝试删除损坏的文件
                if os.path.exists(self.concepts_cache):
                    try:
                        if not self._is_excel_file_valid(self.concepts_cache):
                            backup_file = self.concepts_cache.replace('.xlsx', '_backup.xlsx')
                            try:
                                os.rename(self.concepts_cache, backup_file)
                            except:
                                pass
                    except:
                        pass
    
    def get_stock_list(self, force_refresh: bool = False) -> Optional[pd.DataFrame]:
        """
        从缓存获取股票列表
        Args:
            force_refresh: 是否强制刷新
        Returns:
            股票列表DataFrame
        """
        if not force_refresh and self._is_cache_valid(self.stock_list_cache, 'stock_list'):
            try:
                df = pd.read_excel(self.stock_list_cache, engine='openpyxl')
                # 确保代码是字符串类型，并格式化为6位（补零）
                if 'code' in df.columns:
                    df['code'] = df['code'].astype(str).str.zfill(6)
                return df
            except Exception as e:
                print(f"读取股票列表缓存失败: {e}")
        return None
    
    def save_stock_list(self, stock_list: pd.DataFrame):
        """
        保存股票列表到缓存
        Args:
            stock_list: 股票列表DataFrame
        """
        try:
            # 确保代码是字符串类型，并格式化为6位（补零）
            stock_list = stock_list.copy()
            if 'code' in stock_list.columns:
                stock_list['code'] = stock_list['code'].astype(str).str.zfill(6)
            stock_list.to_excel(self.stock_list_cache, index=False, engine='openpyxl')
        except Exception as e:
            print(f"保存股票列表缓存失败: {e}")
    
    def get_kline(self, symbol: str, cache_type: str = 'stock', 
                  period: str = 'daily', force_refresh: bool = False) -> Optional[pd.DataFrame]:
        """
        从缓存获取K线数据（智能检查最新交易日数据）
        Args:
            symbol: 股票代码/板块名称/概念名称
            cache_type: 缓存类型 ('stock', 'sector', 'concept')
            period: 周期 ('daily', 'weekly', 'monthly')
            force_refresh: 是否强制刷新
        Returns:
            K线数据DataFrame
        """
        if force_refresh:
            return None
        
        cache_file = self._get_kline_cache_path(symbol, cache_type, period)
        if self._is_cache_valid(cache_file, 'kline'):
            try:
                df = pd.read_excel(cache_file, engine='openpyxl')
                # 确保date列是datetime类型（处理日期格式）
                if 'date' in df.columns:
                    df['date'] = pd.to_datetime(df['date']).dt.date
                    df['date'] = pd.to_datetime(df['date'])
                    # 检查最新交易日数据是否存在
                    latest_date = df['date'].max()
                    today = datetime.now().date()
                    # 如果最新数据是今天或昨天（考虑交易日），则可以使用缓存
                    if isinstance(latest_date, pd.Timestamp):
                        latest_date_date = latest_date.date()
                    else:
                        latest_date_date = latest_date
                    if latest_date_date >= today - timedelta(days=1):
                        return df
                    # 否则返回None，需要重新下载
                else:
                    return df
            except Exception as e:
                print(f"读取K线缓存失败 ({symbol}): {e}")
        return None
    
    def has_latest_trading_day_data(self, symbol: str, cache_type: str = 'stock', 
                                    period: str = 'daily') -> bool:
        """
        检查是否有最新交易日的数据（用于智能跳过下载）
        Args:
            symbol: 股票代码/板块名称/概念名称
            cache_type: 缓存类型 ('stock', 'sector', 'concept')
            period: 周期 ('daily', 'weekly', 'monthly')
        Returns:
            是否有最新交易日数据
        """
        cache_file = self._get_kline_cache_path(symbol, cache_type, period)
        if not os.path.exists(cache_file):
            return False
        
        try:
            df = pd.read_excel(cache_file, engine='openpyxl')
            if df.empty:
                return False
            
            if 'date' in df.columns:
                df['date'] = pd.to_datetime(df['date'])
                latest_date = df['date'].max()
                today = datetime.now().date()
                current_time = datetime.now().time()
                weekday = datetime.now().weekday()
                
                # 判断是否应该使用昨天的数据
                # 如果当前在交易时间内，使用昨天的数据（今天数据不完整）
                use_yesterday = False
                if weekday >= 5:  # 周末
                    use_yesterday = True
                elif current_time < datetime.strptime('09:30', '%H:%M').time():  # 还没开盘
                    use_yesterday = True
                elif (datetime.strptime('09:30', '%H:%M').time() <= current_time <= datetime.strptime('11:30', '%H:%M').time()) or \
                     (datetime.strptime('13:00', '%H:%M').time() <= current_time <= datetime.strptime('15:00', '%H:%M').time()):  # 交易中
                    use_yesterday = True
                elif datetime.strptime('11:30', '%H:%M').time() < current_time < datetime.strptime('13:00', '%H:%M').time():  # 午休
                    use_yesterday = True
                
                # 如果应该使用昨天的数据，检查是否有昨天的数据
                if use_yesterday:
                    # 需要昨天或更早的完整数据
                    yesterday = today - timedelta(days=1)
                    # 确保不是周末
                    while yesterday.weekday() >= 5:
                        yesterday = yesterday - timedelta(days=1)
                    return latest_date.date() >= yesterday
                else:
                    # 可以使用今天的数据（收盘后）
                    return latest_date.date() >= today - timedelta(days=1)
            else:
                # 如果没有date列，检查文件修改时间
                return self._is_cache_valid(cache_file, 'kline')
        except:
            return False
    
    def save_kline(self, symbol: str, data: pd.DataFrame, 
                   cache_type: str = 'stock', period: str = 'daily', 
                   incremental: bool = True):
        """
        保存K线数据到缓存（支持增量更新和自动清理）
        Args:
            symbol: 股票代码/板块名称/概念名称
            data: K线数据DataFrame
            cache_type: 缓存类型 ('stock', 'sector', 'concept')
            period: 周期 ('daily', 'weekly', 'monthly')
            incremental: 是否增量更新（默认True，会合并现有缓存数据）
        """
        try:
            if data is None or data.empty:
                return
            
            cache_file = self._get_kline_cache_path(symbol, cache_type, period)
            
            # 确保date列存在并转换为日期类型（不包含时间）
            if 'date' in data.columns:
                data_to_save = data.copy()
                # 将日期转换为只包含日期部分（不包含时间）
                data_to_save['date'] = pd.to_datetime(data_to_save['date']).dt.date
            else:
                data_to_save = data.copy()
                if '日期' in data.columns:
                    data_to_save['date'] = pd.to_datetime(data['日期']).dt.date
            
            # 增量更新：合并现有缓存数据
            if incremental and os.path.exists(cache_file):
                try:
                    existing_df = pd.read_excel(cache_file, engine='openpyxl')
                    if not existing_df.empty and 'date' in existing_df.columns:
                        # 确保date列是日期类型（不包含时间）
                        existing_df['date'] = pd.to_datetime(existing_df['date']).dt.date
                        data_to_save['date'] = pd.to_datetime(data_to_save['date']).dt.date if isinstance(data_to_save['date'].iloc[0], str) else data_to_save['date']
                        
                        # 合并数据：保留旧数据，用新数据更新或追加
                        # 移除旧数据中与新数据日期重复的记录
                        existing_df = existing_df[~existing_df['date'].isin(data_to_save['date'])]
                        
                        # 合并新旧数据
                        data_to_save = pd.concat([existing_df, data_to_save], ignore_index=True)
                        
                        # 按日期排序并去重（保留最新的）
                        data_to_save = data_to_save.sort_values('date').drop_duplicates(subset=['date'], keep='last')
                except Exception as e:
                    # 如果读取现有缓存失败，直接使用新数据
                    print(f"读取现有K线缓存失败，使用新数据覆盖 ({symbol}): {e}")
            
            # 数据清理：只保留最近N天的数据（可配置）
            if 'date' in data_to_save.columns and not data_to_save.empty:
                retention_days = getattr(config, 'KLINE_CACHE_RETENTION_DAYS', 250)
                # 确保date列是datetime类型
                data_to_save['date'] = pd.to_datetime(data_to_save['date'])
                # 获取最新日期
                latest_date = data_to_save['date'].max()
                if isinstance(latest_date, pd.Timestamp):
                    latest_date = latest_date.date()
                elif hasattr(latest_date, 'date'):
                    latest_date = latest_date.date()
                else:
                    latest_date = pd.to_datetime(latest_date).date()
                
                # 计算保留日期范围
                cutoff_date = latest_date - timedelta(days=retention_days)
                
                # 过滤数据：只保留最近N天的数据
                data_to_save = data_to_save[data_to_save['date'].dt.date >= cutoff_date].copy()
                
                # 将日期转换回日期格式（不包含时间）用于保存
                data_to_save['date'] = data_to_save['date'].dt.date
            
            # 保存到缓存（使用openpyxl引擎，确保日期格式正确）
            if not data_to_save.empty:
                # 先保存为Excel
                data_to_save.to_excel(cache_file, index=False, engine='openpyxl')
                
                # 设置日期列的格式为日期（不包含时间）
                try:
                    wb = load_workbook(cache_file)
                    ws = wb.active
                    # 找到date列的索引
                    if 'date' in data_to_save.columns:
                        date_col_idx = list(data_to_save.columns).index('date') + 1
                        # 设置日期格式（只显示日期，不显示时间）
                        date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
                        for row in range(2, ws.max_row + 1):  # 跳过标题行
                            cell = ws.cell(row=row, column=date_col_idx)
                            if cell.value:
                                # 确保单元格格式为日期
                                cell.number_format = 'YYYY-MM-DD'
                    wb.save(cache_file)
                    wb.close()
                except Exception as e:
                    # 如果格式化失败，不影响数据保存
                    pass
        except Exception as e:
            print(f"保存K线缓存失败 ({symbol}): {e}")
    
    def _get_kline_cache_path(self, symbol: str, cache_type: str, period: str) -> str:
        """
        获取K线缓存文件路径
        Args:
            symbol: 股票代码/板块名称/概念名称
            cache_type: 缓存类型 ('stock', 'sector', 'concept')
            period: 周期 ('daily', 'weekly', 'monthly')
        Returns:
            缓存文件路径
        """
        # 清理symbol中的特殊字符，用于文件名
        safe_symbol = str(symbol).replace('/', '_').replace('\\', '_').replace(':', '_')
        
        # 所有K线数据统一放在kline目录下，通过文件名前缀区分类型
        if cache_type == 'stock':
            filename = f"stock_{safe_symbol}_{period}.xlsx"
        elif cache_type == 'sector':
            filename = f"sector_{safe_symbol}_{period}.xlsx"
        elif cache_type == 'concept':
            filename = f"concept_{safe_symbol}_{period}.xlsx"
        else:
            filename = f"{cache_type}_{safe_symbol}_{period}.xlsx"
        
        return os.path.join(self.kline_cache_dir, filename)
    
    def check_cache_completeness(self, stock_codes: List[str], 
                                data_types: List[str] = None) -> Dict[str, Dict]:
        """
        检查缓存完整性，判断是否需要预加载
        Args:
            stock_codes: 股票代码列表
            data_types: 要检查的数据类型列表，如 ['fundamental', 'financial']，None表示检查所有
        Returns:
            缓存完整性统计字典，包含每个类型的覆盖率
        """
        if data_types is None:
            data_types = ['fundamental', 'financial']
        
        if not stock_codes:
            return {dt: {'total': 0, 'cached': 0, 'coverage': 0.0, 'needs_preload': False} for dt in data_types}
        
        # 抽样检查（最多检查前1000只，避免太慢）
        sample_size = min(1000, len(stock_codes))
        sample_codes = stock_codes[:sample_size]
        
        completeness = {}
        for data_type in data_types:
            cached_count = 0
            for code in sample_codes:
                if data_type == 'fundamental':
                    if self.get_fundamental(code, force_refresh=False) is not None:
                        cached_count += 1
                elif data_type == 'financial':
                    if self.get_financial(code, force_refresh=False) is not None:
                        cached_count += 1
            
            coverage = cached_count / sample_size if sample_size > 0 else 0.0
            completeness[data_type] = {
                'total': sample_size,
                'cached': cached_count,
                'coverage': coverage,
                'needs_preload': coverage < 0.5  # 覆盖率低于50%需要预加载
            }
        
        return completeness
    
    def clear_cache(self, cache_type: str = None):
        """
        清除缓存
        Args:
            cache_type: 缓存类型，None表示清除所有
                       'kline': 清除K线缓存
                       'fundamental': 清除基本面缓存
                       'financial': 清除财务缓存
                       'sectors': 清除板块缓存
                       'concepts': 清除概念缓存
                       'stock_list': 清除股票列表缓存
        """
        if cache_type is None:
            # 清除所有缓存
            import shutil
            try:
                # 清除K线缓存
                if os.path.exists(self.kline_cache_dir):
                    for filename in os.listdir(self.kline_cache_dir):
                        file_path = os.path.join(self.kline_cache_dir, filename)
                        try:
                            if os.path.isfile(file_path):
                                os.remove(file_path)
                        except:
                            pass
                
                # 清除meta目录下的缓存文件
                if os.path.exists(self.meta_dir):
                    for filename in os.listdir(self.meta_dir):
                        if filename.endswith('.xlsx'):
                            file_path = os.path.join(self.meta_dir, filename)
                            try:
                                os.remove(file_path)
                            except:
                                pass
                
                print("已清除所有缓存")
            except Exception as e:
                print(f"清除缓存失败: {e}")
        elif cache_type == 'kline':
            # 清除K线缓存
            try:
                if os.path.exists(self.kline_cache_dir):
                    for filename in os.listdir(self.kline_cache_dir):
                        file_path = os.path.join(self.kline_cache_dir, filename)
                        try:
                            if os.path.isfile(file_path):
                                os.remove(file_path)
                        except:
                            pass
                print("已清除K线缓存")
            except Exception as e:
                print(f"清除K线缓存失败: {e}")
        elif cache_type == 'fundamental':
            # 清除基本面缓存
            try:
                if os.path.exists(self.fundamental_cache):
                    os.remove(self.fundamental_cache)
                print("已清除基本面缓存")
            except Exception as e:
                print(f"清除基本面缓存失败: {e}")
        elif cache_type == 'financial':
            # 清除财务缓存
            try:
                if os.path.exists(self.financial_cache):
                    os.remove(self.financial_cache)
                print("已清除财务缓存")
            except Exception as e:
                print(f"清除财务缓存失败: {e}")
        elif cache_type == 'sectors':
            # 清除板块缓存
            try:
                if os.path.exists(self.sectors_cache):
                    os.remove(self.sectors_cache)
                print("已清除板块缓存")
            except Exception as e:
                print(f"清除板块缓存失败: {e}")
        elif cache_type == 'concepts':
            # 清除概念缓存
            try:
                if os.path.exists(self.concepts_cache):
                    os.remove(self.concepts_cache)
                print("已清除概念缓存")
            except Exception as e:
                print(f"清除概念缓存失败: {e}")
        elif cache_type == 'stock_list':
            # 清除股票列表缓存
            try:
                if os.path.exists(self.stock_list_cache):
                    os.remove(self.stock_list_cache)
                print("已清除股票列表缓存")
            except Exception as e:
                print(f"清除股票列表缓存失败: {e}")
        else:
            print(f"未知的缓存类型: {cache_type}")

